import sys, os, time
import urllib.request
import getpass
import py2exe
import requests
import openpyxl
import time
from random import *
import random,string
from selenium import webdriver
from boltons import iterutils
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
path = input("Please enter file path: ")
wk = openpyxl.load_workbook(path)
sh = wk['Sheet1']
rows = sh.max_row
User_email = input("Your Email: ")
password = getpass.getpass("Your Password: ")
#User_password = input("Your Password: ")
CSV_path1 = input("Please enter location to save file: ")
CSV_path =  CSV_path1.replace('/','//')
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()
driver.maximize_window()
driver.get('https://backend.goto.com.pk/catalog/products/index')
time.sleep(5)
Email = driver.find_element_by_id('usersform-email').send_keys(User_email)
Password = driver.find_element_by_id('usersform-password').send_keys(password)
Enter = driver.find_element_by_id('usersform-password').send_keys( u'\ue007')
time.sleep(2)
data = []
for i in range (1, rows+1):
         row_values = []
         SKU_value = sh.cell(i,1).value
         val = iterutils.chunked(SKU_value, 128)
         print(val)
         col1 = sh.max_column
         print(col1)
         catalog_search = driver.find_element_by_xpath('/html/body/div[1]/div/section[1]/div/div[1]/a').click()
         b=1
         while(b==1):
             try:
                 Clear_value = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').clear()
                 Enter_SKU = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').send_keys(val)
                 Enter_SKU_Search = driver.find_element_by_xpath('//*[@id="searchModal"]/div/div/div[2]/div/div[3]/div/input').send_keys(u'\ue007')
                 b=2
             except:
                 pass
         time.sleep(5)
         edit_click = driver.find_element_by_xpath('//*[@id="w0"]/div[2]/div/table/tbody/tr/td[8]/a[@title="Update Product"]').click()
         time.sleep(2)
         linked_prod = driver.find_element_by_xpath('//*[@id="w0"]/div[1]/div/div/ul/li[9]/a[contains(.,"Linked")]').click()
         k=0
         for j in range (2, col1+1):
             SKU_value2 = sh.cell(i,j).value
             print(SKU_value2)
             row_values.append(SKU_value2)
             if SKU_value2 != None:
                 print(j)
                 print(k)
                 val2 = iterutils.chunked(SKU_value2, 128)
                 Clear_clear = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr/td[1]/div/input[@id="dynamicmodel-childskus-%d"]' % (k,)).clear()
                 custom_input = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr/td[1]/div/input[@id="dynamicmodel-childskus-%d"]' % (k,)).send_keys(val2)
                 time.sleep(2)
                 custom_add = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr/td[@class="list-cell__button"]').click()
                 time.sleep(2)
                 k=k+1
         str1 = ", ".join(str(x) for x in row_values)
         values_str = str1.replace(', None','')
         final_array = values_str.split(',')
         print(values_str)
         print(final_array)
         arr_len = str(len(final_array)+1)
         print(arr_len)
         field_cut = driver.find_element_by_xpath('//*[@id="w1"]/table/tbody/tr['+arr_len+']/td[2]').click()
         time.sleep(1)
         custom_submit = driver.find_element_by_xpath('//*[@id="w0"]/div[2]/button').click()
         time.sleep(2)    
         a=1
         while(a==1):
             try:
                driver.find_element_by_xpath('/html/body/div/div/section[2]/div[contains(@class,"alert-success")]')
                a=2
                time.sleep(3)
                product_section = driver.find_element_by_xpath('/html/body/div/aside/section/ul/li[4]/ul/li[4]/a').click()
             except:
                 pass    
         data.append((val))
import pandas as pd
df = pd.DataFrame(data,columns =['Custom-Done SKU'])
df.to_csv(CSV_path+'//DATA-CUSTOM.csv',index=False,encoding='utf-8')
username_drop = driver.find_element_by_xpath('/html/body/div[1]/header/nav/div/ul/li[@class="dropdown user user-menu"]/a').click()
userlogout = driver.find_element_by_xpath('/html/body/div[1]/header/nav/div/ul/li/ul/li[3]/a').click()
time.sleep(2)
driver.close()
         
